from openpyxl import load_workbook
from math import sqrt
import os, fnmatch
import PySimpleGUI as sg
import requests # to get image from the web
import shutil # to save it locally
import csv

r = 56
g = 32
b = 167
filename = 'C:/Users/matte/PycharmProjects/EmbroideryRGB/'

wilcomtch_dir = "C:\Program Files\Wilcom\EmbroideryStudio_e4.5\RES\ThreadCharts"
tch = []

webhexurl = "https://www.colorhexa.com/"


def main(r, g, b):
    extension = os.path.splitext(filename)[1]

    threadList = []

    if extension == '.tch':
        with open(filename, newline='') as f:
            reader = csv.reader(f)
            for c in reader:
                threadList.append([c[0], c[2], int(c[4]), int(c[5]), int(c[6])])

    if extension == '.xlsx':
        wb = load_workbook(filename)
        ws = wb['RawInput']
        for row in ws.iter_rows(max_col=1):
            # threadList.append(row.value)
            for cell in row:
                if not cell.value in [None, '']:
                    c = cell.value.split(',')
                    # print(c[2])
                    threadList.append([c[0], c[2], int(c[4]), int(c[5]), int(c[6])])

    color_diffs = []
    for c in threadList:
        # print(c)
        color_diff = sqrt(abs(r - c[2]) ** 2 + abs(g - c[3]) ** 2 + abs(b - c[4]) ** 2)
        color_diffs.append((color_diff, c[0], c[1], c[2], c[3], c[4]))
    #print(color_diffs)
    color_diffs.sort()
    return min(color_diffs)

def rgb_to_hex(rgb):
    return '%02x%02x%02x' % rgb

def downloadImage(input, output):
    image_url = "https://www.colorhexa.com/" + rgb_to_hex(input) + ".png"
    r = requests.get(image_url, stream=True)

    if r.status_code == 200:
        r.raw.decode_content = True
        with open('input.png', 'wb') as f:
            shutil.copyfileobj(r.raw, f)

    image_url = "https://www.colorhexa.com/" + rgb_to_hex(output) + ".png"
    r = requests.get(image_url, stream=True)

    if r.status_code == 200:
        r.raw.decode_content = True
        with open('output.png', 'wb') as f:
            shutil.copyfileobj(r.raw, f)

if __name__ == '__main__':
    # gettchfiles()
    # loadtch()
    #print(main(178, 232, 1))
    input_img = sg.Image(size=(100, 100), key='input-img')
    output_img = sg.Image(size=(100, 100), key='output-img')
    thread_chart = sg.Text(key='chart', size=(28,1))
    thread_code = sg.Text(key='code', size=(10,1))
    thread_name = sg.Text(key='name', size=(28,1))
    thread_tr = sg.Text(key='tr', size=(10,1))
    thread_tg = sg.Text(key='tg', size=(10,1))
    thread_tb = sg.Text(key='tb', size=(10,1))

    left_pane = [
        #[sg.Image(size=(300, 100), filename='Logo-White.png')],
        [sg.Text("Choose a file: ")],
        [sg.FileBrowse(key='file',initial_folder=wilcomtch_dir, file_types=(("*.xlsx", "*.xlsx"),("*.tch", "*.tch"),))],
        [thread_chart],
        [sg.HSeparator()],
        [sg.Text('Color to Match')],
        [sg.Text('R'), sg.Input(key='r')],
        [sg.Text('G'), sg.Input(key='g')],
        [sg.Text('B'), sg.Input(key='b')],
        [sg.Button('Read'), sg.Exit()]
    ]

    right_pane = [
        [sg.Text('Color to Match')],
        [input_img],
        [sg.HSeparator()],
        [sg.Text('Matched Color')],
        [output_img],
        [sg.Text('Thread Code'), sg.Text('Thread Name')],
        [thread_code, thread_name],
        [sg.Text('R'),thread_tr],
        [sg.Text('G'), thread_tg],
        [sg.Text('B'),thread_tb],
        []
    ]

    row = [
        [
            sg.B('OK'),
            sg.B('Cancel')
        ]
    ]

    layout = [
        [
            sg.Column(left_pane),
            sg.VSeparator(),
            sg.Column(right_pane),
        ]
    ]

    # Create the window
    window = sg.Window("TheEmbroideryNerd - EmbroideryThread Converter", layout)
    # Create an event loop
    while True:
        event, values = window.read()

        filename = values['file']
        r = int(values['r'])
        g = int(values['g'])
        b = int(values['b'])
        output = main(r, g, b)

        print(output)
        downloadImage((r,g,b), (int(output[3]), int(output[4]), int(output[5])))
        output_img.Update(size=(100, 100), filename='output.png')
        input_img.Update(size=(100, 100), filename='input.png')

        window.FindElement('code').Update(output[1])
        window.FindElement('chart').Update(os.path.basename(filename))
        window.FindElement('name').Update(output[2])
        window.FindElement('tr').Update(output[3])
        window.FindElement('tg').Update(output[4])
        window.FindElement('tb').Update(output[5])

        if event == sg.WIN_CLOSED or event == 'Exit':
            break

window.close()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
